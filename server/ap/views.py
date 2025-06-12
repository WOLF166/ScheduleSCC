import json
import datetime

from django.shortcuts import render, redirect
from django.http import HttpResponse, HttpResponseNotFound, JsonResponse
# disp
from openpyxl import load_workbook

from .models import Dispatcher
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from .models import Teacher, Group, Subject, Schedule


def page(request):
    return render(request, "index.html")


def schadd(request, id, groupId, teacherId, subjectId,
           day, week, room, startTime, endTime):

    shed = Schedule.objects.create(id=id,
                                   groupId=groupId,
                                   teacherId=teacherId,
                                   subjectId=subjectId,
                                   day=day,
                                   week=week,
                                   room=room,
                                   startTime=startTime,
                                   endTime=endTime)

    return HttpResponse(shed)


def gradd(request, id, name):

    group = Group.objects.create(id=id, name=name)

    return HttpResponse(group)


def subjadd(request, id, name):

    subject = Subject.objects.create(id=id, name=name)

    return HttpResponse(subject)


def techadd(request, id, name):

    teacher = Teacher.objects.create(id=id, name=name)

    return HttpResponse(teacher)


 # GET
def getSchedule(request):
    group_id = request.GET.get("group")
    if not group_id:
        return HttpResponseNotFound("Не указан параметр group")

    try:
        group = Group.objects.get(id=group_id)
    except Group.DoesNotExist:
        return HttpResponseNotFound("Группа не найдена")

    # Фильтруем расписание по группе
    schedules = Schedule.objects.filter(groupId=group)

    lst = []
    for shed in schedules:
        lst.append({
            "id": shed.id,
            "groupId": shed.groupId.id,
            "group": shed.groupId.name,
            "teacherId": shed.teacherId.id,
            "teacher": shed.teacherId.name,
            "subject": shed.subjectId.name,
            "day": shed.day.isoformat(),
            "week": shed.week,
            "room": shed.room,
            "startTime": shed.startTime,
            "endTime": shed.endTime,
        })

    if lst:
        return JsonResponse(lst, safe=False)
    else:
        return HttpResponseNotFound("Нет расписания")


def getScheduleAll(request):  # Расписание
    schedules = Schedule.objects.all()

    lst = []

    for shed in schedules:
        lst.append({"id": f"{shed.id}",
                    "group": f"{Group.objects.get(id=shed.groupId.pk).name}",
                    "teacher": f"{Teacher.objects.get(id=shed.teacherId.pk).name}",
                    "subject": f"{Subject.objects.get(id=shed.subjectId.pk).name}",
                    "day": f"{shed.day}",
                    "week": f"{shed.week}",
                    "room": f"{shed.room}",
                    "startTime": f"{shed.startTime}",
                    "endTime": f"{shed.endTime}"})

    if lst:
        return HttpResponse(json.dumps(lst), content_type="application/json")
    else:
        return HttpResponseNotFound("Нет расписания")




def getScheduleTeachers(request):  # Расписание
    teachername = request.GET.get("teacher")
    teacher = Teacher.objects.get(name=teachername)
    schedules = Schedule.objects.filter(groupId=teacher.id)

    lst = []

    for shed in schedules:
        lst.append({"id": f"{shed.id}",
                    "groupId": f"{shed.groupId}",
                    "teacherId": f"{shed.teacherId}",
                    "subjectId": f"{shed.subjectId}",
                    "day": f"{shed.day}",
                    "week": f"{shed.week}",
                    "room": f"{shed.room}",
                    "startTime": f"{shed.startTime}",
                    "endTime": f"{shed.endTime}"})
    if lst:
        return HttpResponse(json.dumps(lst), content_type="application/json")
    else:
        return HttpResponseNotFound("Нет преподавателей")


def getAllGroup(request):  # Расписание
    group = Group.objects.all()

    lst = []

    for gr in group:
        lst.append({"id": f"{gr.id}",
                    "name": f"{gr.name}"})
    if lst:
        return HttpResponse(json.dumps(lst), content_type="application/json")
    else:
        return HttpResponseNotFound("Нет групп")


def getAllTeachers(request):  # Расписание
    teacher = Teacher.objects.all()

    lst = []

    for tch in teacher:
        lst.append({"id": f"{tch.id}",
                    "name": f"{tch.name}"})
    if lst:
        return HttpResponse(json.dumps(lst), content_type="application/json")
    else:
        return HttpResponseNotFound("Нет преподавателей")


def getAllSubjects(request):  # Расписание
    subject = Subject.objects.all()

    lst = []

    for sbj in subject:
        lst.append({"id": f"{sbj.id}",
                    "name": f"{sbj.name}"})
    if lst:
        return HttpResponse(json.dumps(lst), content_type="application/json")
    else:
        return HttpResponseNotFound("Нет групп")
# конец api


def mainSchedule(request):
    return render(request, "page.html")


@require_http_methods(["GET"])
def get_schedule(request):
    year = request.GET.get('year', datetime.date.today().year)
    week = request.GET.get('week', datetime.date.today().isocalendar()[1])
    group_id = request.GET.get('group')
    teacher_id = request.GET.get('teacher')

    queryset = Schedule.objects.filter(week=week)

    if year:
        queryset = queryset.filter(day__year=year)
    if group_id:
        queryset = queryset.filter(groupId=group_id)
    if teacher_id:
        queryset = queryset.filter(teacherId=teacher_id)

    data = list(queryset.values(
        'id', 'day', 'week', 'startTime', 'endTime', 'room',
        'groupId__name', 'teacherId__name', 'subjectId__name'
    ).order_by('day', 'startTime'))

    # Переименовываем поля для фронтенда
    for item in data:
        item['group'] = item.pop('groupId__name')
        item['teacher'] = item.pop('teacherId__name')
        item['subject'] = item.pop('subjectId__name')
        item['day'] = item['day'].isoformat()

    return JsonResponse(data, safe=False)


def dispatcher_login_view(request):
    if request.method == "POST":
        login_input = request.POST.get("login")
        password_input = request.POST.get("password")
        try:
            dispatcher = Dispatcher.objects.get(login=login_input, password=password_input)
            response = redirect("dispatcher_dashboard")
            response.set_cookie("dispatcher_id", dispatcher.id, httponly=True)
            return response
        except Dispatcher.DoesNotExist:
            return render(request, "login.html", {"error": "Неверный логин или пароль"})
    return render(request, "login.html")


def dispatcher_dashboard_view(request):
    dispatcher_id = request.COOKIES.get("dispatcher_id")
    if dispatcher_id and Dispatcher.objects.filter(id=dispatcher_id).exists():
        return render(request, "dispatcher_dashboard.html")
    return redirect("dispatcher_login")


def dispatcher_logout_view(request):
    response = redirect("dispatcher_login")
    # Delete the dispatcher_id cookie by setting max_age=0
    response.delete_cookie("dispatcher_id")
    return response


def dispatcher_required(view_func):
    def wrapper(request, *args, **kwargs):
        dispatcher_id = request.COOKIES.get("dispatcher_id")
        if not dispatcher_id:
            return JsonResponse({"error": "Unauthorized"}, status=401)
        # Optionally verify dispatcher exists in DB here
        return view_func(request, *args, **kwargs)
    return wrapper

@require_http_methods(["GET"])
def teachers_list(request):
    teachers = list(Teacher.objects.values("id", "name"))
    return JsonResponse(teachers, safe=False)

@require_http_methods(["POST"])
@csrf_exempt
@dispatcher_required
def add_teacher(request):
    data = json.loads(request.body)
    name = data.get("name")
    if not name:
        return JsonResponse({"error": "Name required"}, status=400)
    teacher = Teacher.objects.create(name=name)
    return JsonResponse({"id": teacher.id, "name": teacher.name}, status=201)

@require_http_methods(["GET"])
def groups_list(request):
    groups = list(Group.objects.values("id", "name"))
    return JsonResponse(groups, safe=False)

@require_http_methods(["POST"])
@csrf_exempt
@dispatcher_required
def add_group(request):
    data = json.loads(request.body)
    name = data.get("name")
    if not name:
        return JsonResponse({"error": "Name required"}, status=400)
    group = Group.objects.create(name=name)
    return JsonResponse({"id": group.id, "name": group.name}, status=201)

@require_http_methods(["GET"])
def subjects_list(request):
    subjects = list(Subject.objects.values("id", "name"))
    return JsonResponse(subjects, safe=False)

@require_http_methods(["POST"])
@csrf_exempt
@dispatcher_required
def add_subject(request):
    data = json.loads(request.body)
    name = data.get("name")
    if not name:
        return JsonResponse({"error": "Name required"}, status=400)
    subject = Subject.objects.create(name=name)
    return JsonResponse({"id": subject.id, "name": subject.name}, status=201)

@require_http_methods(["POST"])
@csrf_exempt
@dispatcher_required
def add_schedule(request):
    data = json.loads(request.body)
    try:
        group = Group.objects.get(id=data.get("groupId"))
        teacher = Teacher.objects.get(id=data.get("teacherId"))
        subject = Subject.objects.get(id=data.get("subjectId"))
        day = data.get("day")
        week = data.get("week")
        room = data.get("room")
        start_time = data.get("startTime")
        end_time = data.get("endTime")
    except (Group.DoesNotExist, Teacher.DoesNotExist, Subject.DoesNotExist):
        return JsonResponse({"error": "Invalid foreign key"}, status=400)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=400)

    schedule = Schedule.objects.create(
        groupId=group,
        teacherId=teacher,
        subjectId=subject,
        day=day,
        week=week,
        room=room,
        startTime=start_time,
        endTime=end_time,
    )
    return JsonResponse({"id": schedule.id}, status=201)

# --- Teacher ---

@require_http_methods(["PUT"])
@csrf_exempt
@dispatcher_required
def update_teacher(request, teacher_id):
    try:
        teacher = Teacher.objects.get(id=teacher_id)
    except Teacher.DoesNotExist:
        return JsonResponse({"error": "Teacher not found"}, status=404)

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({"error": "Invalid JSON"}, status=400)

    name = data.get("name")
    if not name:
        return JsonResponse({"error": "Name required"}, status=400)

    teacher.name = name
    teacher.save()
    return JsonResponse({"id": teacher.id, "name": teacher.name}, status=200)


@require_http_methods(["DELETE"])
@csrf_exempt
@dispatcher_required
def delete_teacher(request, teacher_id):
    try:
        teacher = Teacher.objects.get(id=teacher_id)
    except Teacher.DoesNotExist:
        return JsonResponse({"error": "Teacher not found"}, status=404)

    teacher.delete()
    return JsonResponse({"result": "Teacher deleted"}, status=204)


# --- Group ---

@require_http_methods(["PUT"])
@csrf_exempt
@dispatcher_required
def update_group(request, group_id):
    try:
        group = Group.objects.get(id=group_id)
    except Group.DoesNotExist:
        return JsonResponse({"error": "Group not found"}, status=404)

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({"error": "Invalid JSON"}, status=400)

    name = data.get("name")
    if not name:
        return JsonResponse({"error": "Name required"}, status=400)

    group.name = name
    group.save()
    return JsonResponse({"id": group.id, "name": group.name}, status=200)


@require_http_methods(["DELETE"])
@csrf_exempt
@dispatcher_required
def delete_group(request, group_id):
    try:
        group = Group.objects.get(id=group_id)
    except Group.DoesNotExist:
        return JsonResponse({"error": "Group not found"}, status=404)

    group.delete()
    return JsonResponse({"result": "Group deleted"}, status=204)


# --- Subject ---

@require_http_methods(["PUT"])
@csrf_exempt
@dispatcher_required
def update_subject(request, subject_id):
    try:
        subject = Subject.objects.get(id=subject_id)
    except Subject.DoesNotExist:
        return JsonResponse({"error": "Subject not found"}, status=404)

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({"error": "Invalid JSON"}, status=400)

    name = data.get("name")
    if not name:
        return JsonResponse({"error": "Name required"}, status=400)

    subject.name = name
    subject.save()
    return JsonResponse({"id": subject.id, "name": subject.name}, status=200)


@require_http_methods(["DELETE"])
@csrf_exempt
@dispatcher_required
def delete_subject(request, subject_id):
    try:
        subject = Subject.objects.get(id=subject_id)
    except Subject.DoesNotExist:
        return JsonResponse({"error": "Subject not found"}, status=404)

    subject.delete()
    return JsonResponse({"result": "Subject deleted"}, status=204)


@require_http_methods(["PUT"])
@csrf_exempt
@dispatcher_required
def update_schedule(request, schedule_id):
    try:
        schedule = Schedule.objects.get(id=schedule_id)
    except Schedule.DoesNotExist:
        return JsonResponse({"error": "Schedule not found"}, status=404)

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({"error": "Invalid JSON"}, status=400)

    # Проверяем и обновляем поля
    group_id = data.get("groupId")
    teacher_id = data.get("teacherId")
    subject_id = data.get("subjectId")
    day = data.get("day")
    week = data.get("week")
    room = data.get("room")
    start_time = data.get("startTime")
    end_time = data.get("endTime")

    if group_id:
        try:
            group = Group.objects.get(id=group_id)
            schedule.groupId = group
        except Group.DoesNotExist:
            return JsonResponse({"error": "Group not found"}, status=404)

    if teacher_id:
        try:
            teacher = Teacher.objects.get(id=teacher_id)
            schedule.teacherId = teacher
        except Teacher.DoesNotExist:
            return JsonResponse({"error": "Teacher not found"}, status=404)

    if subject_id:
        try:
            subject = Subject.objects.get(id=subject_id)
            schedule.subjectId = subject
        except Subject.DoesNotExist:
            return JsonResponse({"error": "Subject not found"}, status=404)

    if day:
        schedule.day = day

    if week is not None:
        schedule.week = week

    if room is not None:
        schedule.room = room

    if start_time is not None:
        schedule.startTime = start_time

    if end_time is not None:
        schedule.endTime = end_time

    schedule.save()

    return JsonResponse({
        "id": schedule.id,
        "groupId": schedule.groupId.id,
        "teacherId": schedule.teacherId.id,
        "subjectId": schedule.subjectId.id,
        "day": schedule.day.isoformat(),
        "week": schedule.week,
        "room": schedule.room,
        "startTime": schedule.startTime,
        "endTime": schedule.endTime,
    }, status=200)


@require_http_methods(["DELETE"])
@csrf_exempt
@dispatcher_required
def delete_schedule(request, schedule_id):
    try:
        schedule = Schedule.objects.get(id=schedule_id)
    except Schedule.DoesNotExist:
        return JsonResponse({"error": "Schedule not found"}, status=404)

    schedule.delete()
    return JsonResponse({"result": "Schedule deleted"}, status=204)


@require_http_methods(["GET"])
def schedule_list(request):
    schedule = list(Schedule.objects.values("id", "groupId", "teacherId", "subjectId", "day", "week", "room", "startTime", "endTime"))
    return JsonResponse(schedule, safe=False)

# -------------------------------


def parse_weeks(weeks_str):
    """
    Парсит строку с неделями вида "1-3,5,7-9" в список чисел [1,2,3,5,7,8,9]
    """
    weeks = set()
    for part in weeks_str.split(","):
        part = part.strip()
        if "-" in part:
            start, end = part.split("-")
            weeks.update(range(int(start), int(end) + 1))
        elif part.isdigit():
            weeks.add(int(part))
    return sorted(weeks)


def get_date_for_week_and_day(first_date, first_weekday_name, target_week, target_weekday_name):
    """
    Вычисляет дату занятия по дате первого учебного дня, названию первого дня недели,
    номеру недели и названию дня недели занятия.
    """
    days_map = {
        "понедельник": 0,
        "вторник": 1,
        "среда": 2,
        "четверг": 3,
        "пятница": 4,
        "суббота": 5,
        "воскресенье": 6,
    }

    first_weekday = days_map.get(first_weekday_name.lower())
    target_weekday = days_map.get(target_weekday_name.lower())

    if first_weekday is None or target_weekday is None:
        raise ValueError(f"Неизвестный день недели: {first_weekday_name} или {target_weekday_name}")

    # Коррекция даты первого учебного дня до понедельника той недели
    first_monday = first_date - datetime.timedelta(days=first_weekday)

    # Дата занятия = первый понедельник + (неделя-1)*7 + смещение по дню недели
    delta_days = (target_week - 1) * 7 + target_weekday
    lesson_date = first_monday + datetime.timedelta(days=delta_days)
    return lesson_date

@csrf_exempt
@dispatcher_required
def upload_schedule(request):
    if request.method != "POST":
        return JsonResponse({"error": "Только POST-запросы разрешены"}, status=405)

    excel_file = request.FILES.get("excel_file")
    if not excel_file:
        return JsonResponse({"error": "Файл не загружен"}, status=400)

    try:
        wb = load_workbook(filename=excel_file, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))

        # Первая строка: дата первого учебного дня и день недели
        first_date_raw = rows[0][0]
        first_weekday_name = rows[0][1]

        if isinstance(first_date_raw, datetime.datetime):
            first_date = first_date_raw.date()
        elif isinstance(first_date_raw, datetime.date):
            first_date = first_date_raw
        else:
            first_date = datetime.datetime.strptime(str(first_date_raw), "%Y-%m-%d").date()

        added_count = 0
        row_idx = 2  # начинаем со строки с группами

        while row_idx < len(rows):
            row = rows[row_idx]
            if not row or not row[0]:
                row_idx += 1
                continue

            if str(row[0]).startswith("Группа:"):
                group_name = str(row[0]).replace("Группа:", "").strip()
                group = Group.objects.filter(name=group_name).first()
                if not group:
                    print(f"Group not found: '{group_name}'")
                    row_idx += 1
                    continue

                # Пропускаем заголовок таблицы (следующая строка)
                row_idx += 2

                while row_idx < len(rows):
                    data_row = rows[row_idx]
                    if not data_row or not data_row[0]:
                        break  # Конец блока группы

                    try:
                        day_name, pair_num, start, end, subj_name, teacher_name, weeks_str, room = data_row[:8]
                    except Exception:
                        row_idx += 1
                        continue

                    subject = Subject.objects.filter(name__iexact=str(subj_name).strip()).first()
                    if not subject:
                        print(f"Subject not found: '{subj_name}'")
                        row_idx += 1
                        continue

                    teacher = Teacher.objects.filter(name__iexact=str(teacher_name).strip()).first()
                    if not teacher:
                        print(f"Teacher not found: '{teacher_name}'")
                        row_idx += 1
                        continue

                    weeks = parse_weeks(str(weeks_str))

                    for week in weeks:
                        lesson_date = get_date_for_week_and_day(first_date, first_weekday_name, week, day_name)
                        Schedule.objects.create(
                            groupId=group,
                            subjectId=subject,
                            teacherId=teacher,
                            day=lesson_date,
                            week=week,
                            room=room,
                            startTime=start,
                            endTime=end,
                        )
                        added_count += 1

                    row_idx += 1
            else:
                row_idx += 1

        return JsonResponse({"status": "success", "added": added_count})
    except Exception as e:
        return JsonResponse({"error": f"Ошибка обработки файла: {e}"}, status=500)


@csrf_exempt
@require_http_methods(["DELETE"])
@dispatcher_required
def delete_all_schedules(request):
    try:
        Schedule.objects.all().delete()
        return JsonResponse({"status": "success", "message": "Все записи удалены"})
    except Exception as e:
        return JsonResponse({"status": "error", "message": str(e)}, status=500)
